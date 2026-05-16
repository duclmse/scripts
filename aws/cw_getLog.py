#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CloudWatch Log Keyword Search Tool v3

Features:
  - Multiple keywords with AND / OR logic
  - Dynamic log group discovery via glob patterns or a file
  - Per-group output files named with the time range
  - Full stream download with optional context lines around matches
  - Log-level pre-filter (ERROR / WARN / INFO / DEBUG)
  - Progress bars (tqdm, graceful fallback if not installed)
  - Summary sheet per file + combined _summary_ file
  - Keyword highlighting + header styling in Excel
  - First row frozen in every worksheet
  - Dry-run mode (scan only, no file writes)
  - Max-events cap per log group
  - UTC or local timestamps
  - Resume via stream cache (--resume)
  - Throttle-aware retry with exponential back-off
"""

from __future__ import annotations

import argparse
import configparser
import csv
import fnmatch
import json
import os
import random
import re
import sys
import threading
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone as dt_timezone

import boto3
from botocore.config import Config
from botocore.exceptions import ClientError
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from tqdm import tqdm as _tqdm_cls  # type: ignore[import]
    _HAS_TQDM = True
except ImportError:
    _HAS_TQDM = False
    _tqdm_cls = None  # type: ignore[assignment]


# =============================================================================
# DEFAULT LOG GROUPS  — used when neither --log-group nor --log-groups-file
#                       is provided on the command line.
# =============================================================================
DEFAULT_LOG_GROUPS: list[str] = [
    # "/aws/lambda/dev-fms-lambda-corporateProcess",
    # "/ecs/dev-fms-axa-ecs-td-backend",
    # "/ecs/dev-fms-bmw-ecs-td-backend",
    # "/ecs/dev-fms-bus-ecs-td-backend",
    # "/ecs/dev-fms-byd-ecs-td-backend",
    # "/ecs/dev-fms-common-ecs-td-backend",
    # "/ecs/dev-fms-common-ecs-td-frontend",
    # "/ecs/dev-fms-jaa-ecs-td-backend",
    # "/ecs/dev-fms-jlr-ecs-td-backend",
    # "/ecs/dev-fms-ker-ecs-td-backend",
    # "/ecs/dev-fms-nfl-ecs-td-backend",
    # "/ecs/dev-fms-share-ecs-td-backend",
    # "/ecs/dev-fms-srs-ecs-td-backend",
    # "/ecs/dev-fms-ssj-ecs-td-backend",
    # "/ecs/dev-fms-zrj-ecs-td-backend",
    "/ecs/dev-fms-zur-ecs-td-backend",
]

DEFAULT_REGION = "ap-northeast-1"
DEFAULT_CREDENTIALS = r"C:\\Users\\VMO\\.aws\\credentials"

# Concurrency
SCAN_WORKERS = 8
STREAM_WORKERS = 8
MAX_API_RETRIES = 10
_API_SEMAPHORE = threading.Semaphore(16)   # cap total concurrent API calls

# Excel styles
_HDR_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HDR_FONT = Font(color="FFFFFF", bold=True)
_MATCH_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
_CTX_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")


# =============================================================================
# TQDM FALLBACK
# =============================================================================
if _HAS_TQDM:
    tqdm = _tqdm_cls  # type: ignore[misc]
else:
    class tqdm:  # type: ignore[no-redef]
        """Minimal no-op stand-in when tqdm is not installed."""

        def __init__(self, iterable=None, total=None, desc=None, unit=None, **kw):
            self._it = iterable

        def __iter__(self):
            return iter(self._it) if self._it is not None else iter([])

        def __enter__(self):
            return self

        def __exit__(self, *a):
            pass

        def update(self, n: int = 1):
            pass

        def set_postfix(self, **kw):
            pass

        def set_description(self, s: str):
            pass

        def close(self):
            pass

        @staticmethod
        def write(s: str):
            print(s)


# =============================================================================
# UTILITIES
# =============================================================================

def sanitize_filename(s: str) -> str:
    """Replace characters that are illegal in file / folder names."""
    return re.sub(r'[\\/:*?"<>|\s]+', "_", s.strip("/"))


def ts_to_iso(ms: int, utc: bool = False) -> str:
    """Convert epoch-milliseconds to ISO-8601 string."""
    if utc:
        return datetime.fromtimestamp(ms / 1000.0, tz=dt_timezone.utc).isoformat()
    return datetime.fromtimestamp(ms / 1000.0).isoformat()


def parse_time_to_millis(value: str) -> int:
    """Accept epoch-seconds (integer string) or common datetime strings."""
    value = value.strip()
    if value.isdigit():
        return int(value) * 1000
    for fmt in (
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
    ):
        try:
            dt = datetime.strptime(value, fmt)
            return int(dt.timestamp() * 1000)
        except ValueError:
            continue
    raise ValueError(f"Cannot parse time value: {value!r}")


def load_credentials_from_file(path: str, profile: str = "default"):
    if not os.path.exists(path):
        raise FileNotFoundError(f"Credentials file not found: {path}")
    cfg = configparser.RawConfigParser()
    if not cfg.read(path):
        raise RuntimeError(f"Cannot read credentials file: {path}")
    if profile not in cfg.sections():
        raise RuntimeError(f"Profile [{profile}] not found in {path}")
    key_id = cfg.get(profile, "aws_access_key_id", fallback=None)
    secret = cfg.get(profile, "aws_secret_access_key", fallback=None)
    token = cfg.get(profile, "aws_session_token", fallback=None)
    if not key_id or not secret:
        raise RuntimeError(
            f"Profile [{profile}] is missing aws_access_key_id or aws_secret_access_key."
        )
    return key_id, secret, token


# =============================================================================
# KEYWORD / FILTER PATTERN HELPERS
# =============================================================================

def build_filter_pattern(keywords: list[str], logic: str) -> str:
    """
    CloudWatch Logs filter-pattern syntax:
      AND  → '"kw1" "kw2"'       all terms must appear
      OR   → '?"kw1" ?"kw2"'    any term may appear
    """
    escaped = ['"' + k.replace('"', '\\"') + '"' for k in keywords]
    if logic.upper() == "OR":
        return " ".join("?" + e for e in escaped)
    return " ".join(escaped)


def message_matches(message: str, keywords_lower: list[str], logic: str) -> bool:
    msg = message.lower()
    if logic.upper() == "OR":
        return any(k in msg for k in keywords_lower)
    return all(k in msg for k in keywords_lower)


def level_matches(message: str, level: str | None) -> bool:
    if not level:
        return True
    # "WARN" matches both "WARN" and "WARNING"
    return level.upper() in message.upper()


# =============================================================================
# THROTTLE-AWARE API WRAPPER
# =============================================================================

def _api_call(fn, **kwargs):
    """Call an AWS API function, retrying on throttling with exponential back-off."""
    for attempt in range(MAX_API_RETRIES):
        try:
            with _API_SEMAPHORE:
                return fn(**kwargs)
        except ClientError as exc:
            code = exc.response["Error"]["Code"]
            if code in ("ThrottlingException", "ServiceUnavailableException",
                        "RequestLimitExceeded"):
                wait = min(2 ** attempt + random.uniform(0, 1), 30)
                time.sleep(wait)
                continue
            raise
    raise RuntimeError(f"API call exceeded {MAX_API_RETRIES} retries (throttling).")


# =============================================================================
# LOG GROUP DISCOVERY
# =============================================================================

def discover_log_groups(logs_client, patterns: list[str]) -> list[str]:
    """
    Expand glob patterns to real log-group names via describe_log_groups.
    Literal strings (no wildcards) are returned as-is without an API call.
    """
    result: list[str] = []
    seen: set[str] = set()

    for pattern in patterns:
        if any(c in pattern for c in ("*", "?", "[")):
            prefix = re.split(r"[*?\[]", pattern)[0]
            paginator = logs_client.get_paginator("describe_log_groups")
            for page in paginator.paginate(logGroupNamePrefix=prefix or "/"):
                for lg in page.get("logGroups", []):
                    name = lg["logGroupName"]
                    if fnmatch.fnmatch(name, pattern) and name not in seen:
                        result.append(name)
                        seen.add(name)
        else:
            if pattern not in seen:
                result.append(pattern)
                seen.add(pattern)

    return result


# =============================================================================
# SCAN A SINGLE LOG GROUP
# =============================================================================

def scan_log_group(
    logs_client,
    log_group: str,
    start_ms: int,
    end_ms: int,
    keywords_lower: list[str],
    keyword_logic: str,
    filter_pattern: str,
    level_filter: str | None,
    max_events: int,
    use_utc: bool,
) -> dict:
    """
    Use filter_log_events to find matching events in one log group.
    Returns a dict with totals, matched events, and matched stream keys.
    """
    matched_events: list[dict] = []
    matched_streams: set[tuple] = set()
    stream_index: defaultdict[tuple, int] = defaultdict(int)
    total_scanned = 0
    total_matched = 0
    next_token = None

    while True:
        params: dict = {
            "logGroupName": log_group,
            "startTime": start_ms,
            "endTime": end_ms,
            "interleaved": True,
            "limit": 10000,
            "filterPattern": filter_pattern,
        }
        if next_token:
            params["nextToken"] = next_token

        resp = _api_call(logs_client.filter_log_events, **params)
        events = resp.get("events", [])

        for e in events:
            total_scanned += 1
            message = e.get("message", "")

            if not message_matches(message, keywords_lower, keyword_logic):
                continue
            if not level_matches(message, level_filter):
                continue

            stream = e.get("logStreamName", "")
            key = (log_group, stream)
            stream_index[key] += 1
            total_matched += 1
            matched_streams.add(key)

            ts = e["timestamp"]
            matched_events.append(
                {
                    "logGroup": log_group,
                    "logStream": stream,
                    "timestamp_iso": ts_to_iso(ts, use_utc),
                    "timestamp_epoch_ms": ts,
                    "eventId": e.get("eventId", ""),
                    "ingestionTime_iso": (
                        ts_to_iso(e["ingestionTime"], use_utc)
                        if "ingestionTime" in e
                        else ""
                    ),
                    "event_index_in_stream": stream_index[key],
                    "message": message,
                    "match_type": "MATCH",
                }
            )

            if max_events and total_matched >= max_events:
                return _scan_result(
                    log_group, total_scanned, total_matched,
                    matched_events, matched_streams, capped=True,
                )

        next_token = resp.get("nextToken")
        if not next_token:
            break

    return _scan_result(
        log_group, total_scanned, total_matched,
        matched_events, matched_streams, capped=False,
    )


def _scan_result(log_group, scanned, matched, events, streams, capped) -> dict:
    return {
        "log_group": log_group,
        "total_scanned": scanned,
        "total_matched": matched,
        "matched_events": events,
        "matched_streams": [tuple(k) for k in streams],
        "capped": capped,
    }


# =============================================================================
# FETCH FULL STREAM
# =============================================================================

def fetch_full_stream(
    logs_client,
    log_group: str,
    log_stream: str,
    start_ms: int,
    end_ms: int,
    use_utc: bool,
    cache_dir: str | None,
) -> list[dict]:
    """Download all events for one log stream; optionally persist a cache."""
    # --- check cache ---
    cache_file = None
    if cache_dir:
        safe = sanitize_filename(f"{log_group}__{log_stream}")
        cache_file = os.path.join(cache_dir, f"{safe}.json")
        if os.path.exists(cache_file):
            with open(cache_file, encoding="utf-8") as f:
                cached = json.load(f)
            if cached.get("start_ms") == start_ms and cached.get("end_ms") == end_ms:
                return cached["events"]

    # --- fetch from API ---
    events_out: list[dict] = []
    next_token = None

    while True:
        params: dict = {
            "logGroupName": log_group,
            "logStreamNames": [log_stream],
            "startTime": start_ms,
            "endTime": end_ms,
            "limit": 10000,
            "interleaved": True,
        }
        if next_token:
            params["nextToken"] = next_token

        resp = _api_call(logs_client.filter_log_events, **params)
        for e in resp.get("events", []):
            ts = e["timestamp"]
            events_out.append(
                {
                    "logGroup": log_group,
                    "logStream": log_stream,
                    "timestamp_iso": ts_to_iso(ts, use_utc),
                    "timestamp_epoch_ms": ts,
                    "ingestionTime_iso": (
                        ts_to_iso(e["ingestionTime"], use_utc)
                        if "ingestionTime" in e
                        else ""
                    ),
                    "eventId": e.get("eventId", ""),
                    "message": e.get("message", ""),
                }
            )

        next_token = resp.get("nextToken")
        if not next_token:
            break

    # --- write cache ---
    if cache_file:
        os.makedirs(cache_dir, exist_ok=True)  # type: ignore[arg-type]
        with open(cache_file, "w", encoding="utf-8") as f:
            json.dump(
                {"start_ms": start_ms, "end_ms": end_ms, "events": events_out},
                f,
                ensure_ascii=False,
            )

    return events_out


# =============================================================================
# CONTEXT LINES
# =============================================================================

def apply_context_lines(
    matched_events: list[dict],
    streams_data: dict[tuple, list[dict]],
    context_lines: int,
) -> list[dict]:
    """
    For each matched event, inject up to context_lines events before and after
    it (from the full stream data).  Returns a combined list sorted by timestamp.
    """
    if context_lines <= 0:
        return matched_events

    matched_ids = {e["eventId"] for e in matched_events if e.get("eventId")}
    extra: dict[str, dict] = {}

    for (log_group, stream), events in streams_data.items():
        id_to_idx = {
            e["eventId"]: i for i, e in enumerate(events) if e.get("eventId")
        }
        for ev in matched_events:
            if ev["logGroup"] != log_group or ev["logStream"] != stream:
                continue
            eid = ev.get("eventId", "")
            if eid not in id_to_idx:
                continue
            idx = id_to_idx[eid]
            lo = max(0, idx - context_lines)
            hi = min(len(events) - 1, idx + context_lines)
            for ci in range(lo, hi + 1):
                ce = events[ci]
                cid = ce.get("eventId", "")
                if cid in matched_ids or cid in extra:
                    continue
                extra[cid] = {
                    **ce,
                    "match_type": "CONTEXT",
                    "event_index_in_stream": ci + 1,
                }

    combined = matched_events + list(extra.values())
    combined.sort(key=lambda x: x.get("timestamp_epoch_ms", 0))
    return combined


# =============================================================================
# EXCEL HELPERS
# =============================================================================

def _write_header(ws, headers: list[str]):
    """Write a styled, frozen header row."""
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _autofit(ws):
    """Set column widths based on content (capped at 80)."""
    for col in ws.columns:
        max_len = 10
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max_len + 2, 80
        )


# =============================================================================
# WRITE PER-GROUP EXCEL
# =============================================================================

def write_group_excel(
    path: str,
    log_group: str,
    scan_result: dict,
    streams_data: dict[tuple, list[dict]],
    keywords_lower: list[str],
    context_lines: int,
    use_utc: bool,
):
    wb = Workbook()

    # ---- Sheet 1: matched_events ----
    ws_me = wb.active
    ws_me.title = "matched_events"
    _write_header(
        ws_me,
        [
            "logGroup", "logStream", "timestamp_iso", "eventId",
            "ingestionTime_iso", "event_index_in_stream", "match_type", "message",
        ],
    )

    events = scan_result["matched_events"]
    if context_lines > 0:
        events = apply_context_lines(events, streams_data, context_lines)

    for rec in events:
        ws_me.append(
            [
                rec.get("logGroup", ""),
                rec.get("logStream", ""),
                rec.get("timestamp_iso", ""),
                rec.get("eventId", ""),
                rec.get("ingestionTime_iso", ""),
                rec.get("event_index_in_stream", ""),
                rec.get("match_type", "MATCH"),
                rec.get("message", ""),
            ]
        )
        row_idx = ws_me.max_row
        msg_cell = ws_me.cell(row=row_idx, column=8)
        if rec.get("match_type") == "CONTEXT":
            msg_cell.fill = _CTX_FILL
        else:
            msg_cell.fill = _MATCH_FILL

    _autofit(ws_me)

    # ---- Sheet 2: log_streams ----
    ws_ls = wb.create_sheet("log_streams")
    _write_header(
        ws_ls,
        [
            "logGroup", "logStream", "timestamp_iso",
            "ingestionTime_iso", "eventId", "message",
        ],
    )
    for (_lg, _stream), evs in streams_data.items():
        for ev in evs:
            ws_ls.append(
                [
                    ev.get("logGroup", _lg),
                    ev.get("logStream", _stream),
                    ev.get("timestamp_iso", ""),
                    ev.get("ingestionTime_iso", ""),
                    ev.get("eventId", ""),
                    ev.get("message", ""),
                ]
            )
    _autofit(ws_ls)

    # ---- Sheet 3: summary ----
    ws_sum = wb.create_sheet("summary")
    _write_header(ws_sum, ["Key", "Value"])
    rows = [
        ("Log Group", log_group),
        ("Total Scanned", scan_result["total_scanned"]),
        ("Total Matched", scan_result["total_matched"]),
        ("Streams with Matches", len(scan_result["matched_streams"])),
        ("Context Lines", context_lines),
        ("Capped at Max Events", scan_result.get("capped", False)),
        ("Generated At (UTC)", datetime.now(tz=dt_timezone.utc).isoformat()),
    ]
    for row in rows:
        ws_sum.append(list(row))
    _autofit(ws_sum)

    wb.save(path)


# =============================================================================
# WRITE COMBINED SUMMARY EXCEL
# =============================================================================

def write_combined_summary_excel(
    path: str,
    scan_results: list[dict],
    keywords: list[str],
    keyword_logic: str,
    start_ms: int,
    end_ms: int,
    use_utc: bool,
):
    wb = Workbook()

    # ---- Sheet 1: per-group summary ----
    ws = wb.active
    ws.title = "by_log_group"
    _write_header(
        ws,
        ["logGroup", "totalScanned", "totalMatched", "streamsWithMatches", "capped"],
    )
    for r in scan_results:
        ws.append(
            [
                r["log_group"],
                r["total_scanned"],
                r["total_matched"],
                len(r["matched_streams"]),
                r.get("capped", False),
            ]
        )
    _autofit(ws)

    # ---- Sheet 2: run metadata ----
    ws_meta = wb.create_sheet("metadata")
    _write_header(ws_meta, ["Key", "Value"])
    rows = [
        ("Keywords", ", ".join(keywords)),
        ("Keyword Logic", keyword_logic),
        ("Start Time", ts_to_iso(start_ms, use_utc)),
        ("End Time", ts_to_iso(end_ms, use_utc)),
        ("Log Groups Scanned", len(scan_results)),
        ("Total Scanned", sum(r["total_scanned"] for r in scan_results)),
        ("Total Matched", sum(r["total_matched"] for r in scan_results)),
        ("Generated At (UTC)", datetime.now(tz=dt_timezone.utc).isoformat()),
    ]
    for row in rows:
        ws_meta.append(list(row))
    _autofit(ws_meta)

    wb.save(path)


# =============================================================================
# CSV / JSON OUTPUT
# =============================================================================

_CSV_FIELDS = [
    "logGroup", "logStream", "timestamp_iso", "eventId",
    "ingestionTime_iso", "event_index_in_stream", "match_type", "message",
]


def write_group_csv(
    path: str,
    scan_result: dict,
    streams_data: dict,
    context_lines: int,
):
    events = scan_result["matched_events"]
    if context_lines > 0:
        events = apply_context_lines(events, streams_data, context_lines)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=_CSV_FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(events)


def write_group_json(
    path: str,
    scan_result: dict,
    streams_data: dict,
    context_lines: int,
):
    events = scan_result["matched_events"]
    if context_lines > 0:
        events = apply_context_lines(events, streams_data, context_lines)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(
            {
                "log_group": scan_result["log_group"],
                "summary": {
                    "total_scanned": scan_result["total_scanned"],
                    "total_matched": scan_result["total_matched"],
                },
                "events": events,
            },
            f,
            ensure_ascii=False,
            indent=2,
        )


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description=(
            "Search AWS CloudWatch Logs by keyword(s) and export results "
            "to per-log-group Excel / CSV / JSON files."
        )
    )
    parser.add_argument(
        "--keyword", nargs="+", required=True,
        help="One or more keywords.  e.g.: --keyword ERROR timeout",
    )
    parser.add_argument(
        "--keyword-logic", choices=["AND", "OR"], default="AND",
        help="How to combine multiple keywords (default: AND).",
    )
    parser.add_argument(
        "--start-time", default=None,
        help=(
            'Start time, e.g. "2025-12-04T00:00:00" or "2025-12-04".  '
            "If omitted and --end-time is given, defaults to 30 min before end.  "
            "If both are omitted, defaults to 30 min ago."
        ),
    )
    parser.add_argument(
        "--end-time", default=None,
        help=(
            'End time, e.g. "2025-12-04T23:59:59".  '
            "If omitted and --start-time is given, defaults to 30 min after start.  "
            "If both are omitted, defaults to now."
        ),
    )
    parser.add_argument(
        "--log-group", action="append", dest="log_groups", metavar="PATTERN",
        help=(
            "Log group name or glob pattern.  May be given multiple times.  "
            "e.g.: --log-group '/ecs/dev-fms-*'"
        ),
    )
    parser.add_argument(
        "--log-groups-file",
        help="Text file with one log group (or glob pattern) per line.",
    )
    parser.add_argument(
        "--run-dir", default=None,
        help="Output directory (auto-generated from keywords + timestamp if omitted).",
    )
    parser.add_argument(
        "--credentials-file", default=DEFAULT_CREDENTIALS,
        help=f"AWS credentials file path (default: {DEFAULT_CREDENTIALS}).",
    )
    parser.add_argument(
        "--profile", default="default",
        help="Credentials profile (default: default).",
    )
    parser.add_argument(
        "--region", default=DEFAULT_REGION,
        help=f"AWS region (default: {DEFAULT_REGION}).",
    )
    parser.add_argument(
        "--output-format", choices=["xlsx", "csv", "json"], default="xlsx",
        help="Output format for per-group files (default: xlsx).",
    )
    parser.add_argument(
        "--context-lines", type=int, default=0, metavar="N",
        help="N lines of context before/after each match (default: 0).",
    )
    parser.add_argument(
        "--level", choices=["ERROR", "WARN", "INFO", "DEBUG"], default=None,
        help="Also require this log level to appear in the message.",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Scan only; print matched streams, do not download or write files.",
    )
    parser.add_argument(
        "--max-events", type=int, default=0, metavar="N",
        help="Stop collecting after N matched events per log group (0 = unlimited).",
    )
    parser.add_argument(
        "--timezone", choices=["UTC", "local"], default="local",
        help="Timezone for timestamps in output (default: local).",
    )
    parser.add_argument(
        "--resume", action="store_true",
        help="Re-use cached stream downloads from a previous run if available.",
    )

    args = parser.parse_args()
    use_utc = args.timezone.upper() == "UTC"

    # ---- parse time range ----
    _30_min_ms = 30 * 60 * 1000
    try:
        if args.start_time and args.end_time:
            start_ms = parse_time_to_millis(args.start_time)
            end_ms = parse_time_to_millis(args.end_time)
        elif args.start_time and not args.end_time:
            start_ms = parse_time_to_millis(args.start_time)
            end_ms = start_ms + _30_min_ms
        elif args.end_time and not args.start_time:
            end_ms = parse_time_to_millis(args.end_time)
            start_ms = end_ms - _30_min_ms
        else:
            end_ms = int(datetime.now().timestamp() * 1000)
            start_ms = end_ms - _30_min_ms
    except ValueError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)

    if end_ms <= start_ms:
        print("ERROR: --end-time must be later than --start-time.", file=sys.stderr)
        sys.exit(1)

    duration_h = (end_ms - start_ms) / 3_600_000
    if duration_h > 24:
        print(
            f"WARNING: Time range is {duration_h:.1f}h (> 24h). "
            "Scan may be slow and consume many API calls.",
            file=sys.stderr,
        )

    keywords = args.keyword
    keywords_lower = [k.lower() for k in keywords]
    keyword_logic = args.keyword_logic
    filter_pattern = build_filter_pattern(keywords, keyword_logic)

    # ---- credentials ----
    print(f"Reading credentials: {args.credentials_file}  profile={args.profile}")
    try:
        key_id, secret, token = load_credentials_from_file(
            args.credentials_file, args.profile
        )
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)

    session = boto3.Session(
        aws_access_key_id=key_id,
        aws_secret_access_key=secret,
        aws_session_token=token,
        region_name=args.region,
    )
    logs_client = session.client(
        "logs",
        config=Config(retries={"max_attempts": MAX_API_RETRIES, "mode": "standard"}),
    )

    # ---- collect log group names ----
    patterns: list[str] = list(args.log_groups or [])
    if args.log_groups_file:
        with open(args.log_groups_file, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    patterns.append(line)
    if not patterns:
        patterns = DEFAULT_LOG_GROUPS

    needs_glob = any(any(c in p for c in ("*", "?", "[")) for p in patterns)
    if needs_glob:
        print("Discovering log groups matching pattern(s)…")
        log_groups = discover_log_groups(logs_client, patterns)
        print(f"  Found {len(log_groups)} log group(s).")
    else:
        log_groups = patterns

    if not log_groups:
        print("ERROR: No log groups to scan.", file=sys.stderr)
        sys.exit(1)

    # ---- set up directories and filename fragments ----
    now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    kw_slug = "_".join(sanitize_filename(k) for k in keywords)
    run_dir = args.run_dir or f"CW_{kw_slug}_{now_str}"
    os.makedirs(run_dir, exist_ok=True)

    cache_dir = os.path.join(run_dir, "cache") if args.resume else None

    start_str = datetime.fromtimestamp(start_ms / 1000).strftime("%Y%m%dT%H%M%S")
    end_str = datetime.fromtimestamp(end_ms / 1000).strftime("%Y%m%dT%H%M%S")

    # ==========================================================================
    # PHASE 1 — SCAN LOG GROUPS
    # ==========================================================================
    print(
        f"\nScanning {len(log_groups)} log group(s) "
        f"for {keywords} [{keyword_logic}]  "
        f"{args.start_time} → {args.end_time}"
    )

    scan_results: list[dict] = []
    all_matched_streams: set[tuple] = set()

    with tqdm(total=len(log_groups), desc="Scanning", unit="group") as pbar:
        with ThreadPoolExecutor(max_workers=SCAN_WORKERS) as ex:
            future_to_group = {
                ex.submit(
                    scan_log_group,
                    logs_client, lg, start_ms, end_ms,
                    keywords_lower, keyword_logic, filter_pattern,
                    args.level, args.max_events, use_utc,
                ): lg
                for lg in log_groups
            }
            for fut in as_completed(future_to_group):
                lg = future_to_group[fut]
                try:
                    result = fut.result()
                except Exception as exc:
                    tqdm.write(f"  ERROR [{lg}]: {exc}")
                    pbar.update(1)
                    continue

                scan_results.append(result)
                for key in result["matched_streams"]:
                    all_matched_streams.add(tuple(key))

                cap_tag = " [CAPPED]" if result.get("capped") else ""
                tqdm.write(
                    f"  [{lg}]  scanned={result['total_scanned']}"
                    f"  matched={result['total_matched']}{cap_tag}"
                )
                pbar.update(1)

    total_scanned = sum(r["total_scanned"] for r in scan_results)
    total_matched = sum(r["total_matched"] for r in scan_results)
    print("\n=== SCAN SUMMARY ===")
    print(f"  Total scanned : {total_scanned}")
    print(f"  Total matched : {total_matched}")
    print(f"  Streams hit   : {len(all_matched_streams)}")

    if args.dry_run:
        print("\n[dry-run] Matched streams:")
        for lg, s in sorted(all_matched_streams):
            print(f"  [{lg}]  {s}")
        print("\nDone (dry-run — no files written).")
        return

    # ==========================================================================
    # PHASE 2 — DOWNLOAD FULL STREAMS
    # ==========================================================================
    streams_data: dict[tuple, list[dict]] = {}

    if all_matched_streams:
        print(f"\nDownloading {len(all_matched_streams)} stream(s)…")
        with tqdm(total=len(all_matched_streams), desc="Fetching", unit="stream") as pbar:
            with ThreadPoolExecutor(max_workers=STREAM_WORKERS) as ex:
                future_to_key = {
                    ex.submit(
                        fetch_full_stream,
                        logs_client, lg, stream,
                        start_ms, end_ms, use_utc, cache_dir,
                    ): (lg, stream)
                    for lg, stream in all_matched_streams
                }
                for fut in as_completed(future_to_key):
                    key = future_to_key[fut]
                    try:
                        streams_data[key] = fut.result()
                    except Exception as exc:
                        tqdm.write(f"  ERROR fetching {key}: {exc}")
                    pbar.update(1)
    else:
        print("No matched streams — skipping download phase.")

    # ==========================================================================
    # PHASE 3 — WRITE OUTPUT  (one file per log group)
    # ==========================================================================
    fmt = args.output_format
    ext = fmt
    print(f"\nWriting output ({fmt}) → {os.path.abspath(run_dir)}/")

    for scan_result in scan_results:
        lg = scan_result["log_group"]
        safe_lg = sanitize_filename(lg)
        base = f"{safe_lg}_{start_str}_{end_str}"
        out_path = os.path.join(run_dir, f"{base}.{ext}")

        group_streams = {
            (g, s): v for (g, s), v in streams_data.items() if g == lg
        }

        if fmt == "xlsx":
            write_group_excel(
                out_path, lg, scan_result, group_streams,
                keywords_lower, args.context_lines, use_utc,
            )
        elif fmt == "csv":
            write_group_csv(out_path, scan_result, group_streams, args.context_lines)
        elif fmt == "json":
            write_group_json(out_path, scan_result, group_streams, args.context_lines)

        print(f"  {os.path.basename(out_path)}")

    # combined summary (xlsx only)
    if fmt == "xlsx":
        summary_path = os.path.join(run_dir, f"_summary_{start_str}_{end_str}.xlsx")
        write_combined_summary_excel(
            summary_path, scan_results, keywords, keyword_logic,
            start_ms, end_ms, use_utc,
        )
        print(f"  {os.path.basename(summary_path)}  (combined summary)")

    print(f"\nRun directory: {os.path.abspath(run_dir)}")
    print("Done.")


if __name__ == "__main__":
    main()
